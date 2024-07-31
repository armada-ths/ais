"""
A command that generates 2 files for each exhibitor: a specially formatted {comany-name}.txt and a human-readable {company-name}.xlsx

To use:
> python manage.py generate_bills

consider calling
> python manage.py generate_bills -h
to list options
"""

# Constants, that might change in the future?
# Taken from the example php, you can probably change this if you know what you're doing
static = {
    "customer_number": "000000",
    "profit_center": 11,
    "cost_carrier": 11,
    "row_type": 3,  # no clue what it is exactly
    "tax": 0,
    "result_center": 11,
}


import os, locale, codecs, errno
import datetime
from django.core.management import BaseCommand, CommandError
from exhibitors.models import Exhibitor
from fair.models import Fair
from orders.models import Order, Product
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Generates 2 bills: a specially formatted 'bills.txt' and a human-readable 'bills.xlsx'"
    options = (None,)
    last_locale = None

    def setlocale(self):
        """
        Set locale to either locale provided in options or one of the standard ones
        self.resetlocale() should be called once done with locale-specific stuff
        """
        # An important override avoidance
        if not self.last_locale:
            self.last_locale = locale.getlocale()
        locale_updated = False
        if self.options["locale"]:
            locales = [self.options["locale"]]
        else:
            locales = ["sv_SE", "swedish", "sv_SE.UTF-8", "swedish.UTF-8"]
        for loc in locales:
            try:
                locale.setlocale(locale.LC_ALL, loc)
                locale_updated = True
                self.verbose_write("locale set to {}".format(loc))
                break
            except locale.Error:
                pass
        if not locale_updated:
            self.resetlocale()
            raise Exception("Could not set locale!")

    def resetlocale(self):
        """
        Restore previous locale
        should only be called after successfull self.setlocale()
        """
        if self.last_locale:
            locale.setlocale(locale.LC_ALL, self.last_locale)
            self.last_locale = None

    def write(self, item):
        """
        A print() wrapper, Django recommends using self.stdout for writing, so I do
        """
        self.stdout.write(str(item))

    def verbose_write(self, item, required_level=3):
        """
        A write if-verbosity helper function
        """
        if self.options["verbosity"] >= required_level:
            self.write(item)

    def add_arguments(self, parser):
        """
        Add arguments to the command, arguments prepened with '--' are optional.
        Look into 'python argparse' for help
        """
        now = (
            datetime.datetime.now()
        )  # TODO: Set to air year instead of current year. Maybe variable?
        dir_str = "../../../bills_%s/" % str(
            now.year
        )  # adds current year to default directory
        parser.add_argument(
            "--dirname",
            type=str,
            default=dir_str,
            help="specify a custom relative path to a directory for the generated files (default is '../../../bills_YYYY/', where YYYY is current year)",
        )
        parser.add_argument(
            "--locale",
            type=str,
            default=None,
            help="specify a custom locale to be used (default is swedish)",
        )
        parser.add_argument(
            "--contact",
            type=str,
            default="KONTAKTPERSON",
            help="Specify a contact person to be specified in the bill (default is 'KONTAKTPERSON')",
        )

    def create_exhibitor_txt(self, exhibitor):
        """
        A helper function that creates and writes a new bill txt file for a given exhibitor
        """
        try:
            # creates directory if it does not exist
            os.makedirs(self.options["dirname"])
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise
        with codecs.open(
            self.options["dirname"]
            + (exhibitor.company.name).replace("/", "")
            + ".txt",
            "w+",
            "ISO-8859-1",
        ) as txt_file:
            txt_file.write("Rubrik\tTHS Armada Faktura\r\n")
            txt_file.write("Datumformat YYYY - MM - DD\r\n")
            # custommer number
            txt_file.write(
                "Kundfaktura\t\t\t\t{customer_number}\t\t\t\t\t{cost_carrier}\t\t\t{ref}\t\t\t".format(
                    customer_number=static["customer_number"],
                    cost_carrier=static["cost_carrier"],
                    ref=exhibitor.invoice_reference,
                )
            )

            txt_file.write("Armada, 2017-11-21<CR>")
            txt_file.write("Fakturamärkning: <CR>")
            txt_file.write(
                "For questions and feedback, contact armada@ths.kth.se.<CR><CR><CR>"
            )
            txt_file.write("\t{}\t\t\t\t\t\t\t\t\t".format(self.options["contact"]))
            txt_file.write(
                "{0}<CR>{1}<CR>{2}<CR>{3}<CR><CR>\t\t\t\t\t\t\t\t\t\t\t\t".format(
                    exhibitor.company.name,
                    exhibitor.invoice_address,
                    exhibitor.invoice_address_po_box,
                    exhibitor.invoice_address_zip_code,
                )
            )

            for order in Order.objects.filter(exhibitor=exhibitor):
                if order.amount > 0:
                    self.write_order_txt(txt_file, order)
            txt_file.write("\r\nKundfaktura-slut")

    def write_order_txt(self, out, order):
        """
        A helper function that writes a formatted order into provided stream
        """
        # Comments are copied from php example, so that someone who speaks swedish can make sure I didn't mess up
        out.write("\r\nFakturarad\t")
        out.write("{}\t\t".format(static["row_type"]))
        # Antal
        out.write("{}\t".format(order.amount))  # item_amount
        # å pris
        out.write("{}\t\t".format(order.product.price))  # item_price
        # Title
        out.write("{}\t".format(order.product.name))  # item_title
        # konto
        out.write("{}\t\t\t".format(order.product.coa_number))  # rent_account
        # moms
        out.write("{}\t".format(static["tax"]))
        # enhet
        out.write("{}\t\t\t\t".format("st"))  # item_suffix
        # resultatställe
        out.write("{}\t\t\t".format(static["result_center"]))
        # KB
        out.write("{}\t".format(static["cost_carrier"]))
        out.write("{}".format(order.price()))  # total_price

    def create_readable_output(self, options):
        for exhibitor in Exhibitor.objects.filter(fair=Fair.objects.get(current=True)):
            wb = load_workbook(
                os.path.dirname(os.path.realpath(__file__)) + "/invoice_template.xlsx"
            )
            ws = wb.active  # Select fakturaunderlag-mall sheet

            # Assign company info to template
            ws["E5"] = exhibitor.company.name.replace("/", "")
            ws["C8"] = exhibitor.invoice_reference
            ws["C9"] = "Project Manager THS Armada"
            ws["E8"] = exhibitor.invoice_address
            if exhibitor.invoice_address_po_box:
                ws["E9"] = exhibitor.invoice_address_po_box
                ws["E10"] = exhibitor.invoice_address_zip_code
            else:
                ws["E9"] = exhibitor.invoice_address_zip_code
            ws["A16"] = exhibitor.invoice_additional_information

            current_row = 23
            for order in Order.objects.filter(exhibitor=exhibitor):
                ws["A" + str(current_row)] = order.amount
                ws["B" + str(current_row)] = order.product.name
                ws["E" + str(current_row)] = order.product.price
                ws["G" + str(current_row)] = order.product.coa_number
                ws["H" + str(current_row)] = 11  # Resultatställe armada
                ws["J" + str(current_row)] = 0  # Resultatställe armada
                current_row += 1

            wb.save(
                "{}/{}.xlsx".format(
                    options["dirname"], exhibitor.company.name.replace("/", "")
                )
            )

    def handle(self, *args, **options):
        """
        The meat of the command
        """
        # I know this is bad practice, to change options on the fly, but don't like creating a porcessed-option variable
        options["dirname"] = os.path.join(
            os.path.dirname(os.path.realpath(__file__)), options["dirname"]
        )
        self.options = options
        self.verbose_write("Generating files in {dirname} directory".format(**options))
        # A try-with-resources block, this is to make sure we don't change the locale settings of python application
        try:
            self.setlocale()
            generated_file_count = 0
            # Loop through all exhibitors:
            for exhibitor in Exhibitor.objects.filter(
                fair=Fair.objects.get(current=True)
            ):
                try:
                    self.create_exhibitor_txt(exhibitor)
                    generated_file_count += 1
                except Exception as ex:
                    self.stderr.write(
                        "Got the follwoing exception while creating a file for {} exhibitor:\n{}".format(
                            exhibitor, ex
                        )
                    )
            self.create_readable_output(options)
            self.verbose_write("Generated {} files".format(generated_file_count), 1)

        finally:
            self.resetlocale()
