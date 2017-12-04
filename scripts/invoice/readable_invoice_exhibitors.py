from django.contrib.auth.models import User
from exhibitors.models import Exhibitor
from orders.models import Product, ProductType, Order
from companies.models import Company, Contact
from fair.models import Fair
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors, Font, Color

"""
    This script generate a readable invoice xlsx file for each exhibitor.
    It includes general exhibitor information and summary of purchased products.
"""

for exhibitor in Exhibitor.objects.filter(fair=Fair.objects.get(current=True)):
    wb = load_workbook('scripts/invoice/invoice_template.xlsx')
    ws = wb.active # Select fakturaunderlag-mall sheet
    
    # Assign company info to template
    ws['E5'] = exhibitor.company.name
    ws['C8'] = exhibitor.invoice_reference
    ws['C9'] = 'Project Manager THS Armada'
    ws['E8'] = exhibitor.invoice_address
    if exhibitor.invoice_address_po_box:
        ws['E9'] = exhibitor.invoice_address_po_box
        ws['E10'] = exhibitor.invoice_address_zip_code        
    else:
        ws['E9'] = exhibitor.invoice_address_zip_code
    ws['A16'] = exhibitor.invoice_additional_information

    current_row = 23
    for order in Order.objects.filter(exhibitor=exhibitor):
        ws['A' + str(current_row)] = order.amount
        ws['B' + str(current_row)] = order.product.name
        ws['E' + str(current_row)] = order.product.price
        ws['G' + str(current_row)] = order.product.coa_number
        ws['H' + str(current_row)] = 11 # Resultatställe armada
        ws['J' + str(current_row)] = 0 # Resultatställe armada
        current_row += 1


    # Ugly styling hack to reset stylings
    ft = Font(name='Garamond',
                size=19,
                bold=True,
                color='000066')

    wb.save('scripts/invoice/{0}.xlsx'.format(exhibitor.company.name))
    print('Successfully wrote invoices to file')
